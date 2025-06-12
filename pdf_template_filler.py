"""
PDF Template Filler

This script replaces bracketed fields in a PDF document with values from an Excel file.
For example, it will replace [First Name] with "John" based on Excel data.

Requirements:
    - Python 3.6 or higher
    - Required Python packages:
        pip install PyMuPDF==1.25.3     # For PDF handling (also known as fitz)
        pip install openpyxl==3.0.10    # For Excel file handling
    - No external applications required

Usage:
    1. Prepare your files:
       - PDF template: Use bracketed fields like [Field Name] where you want replacements
       - Excel file: First row should contain headers matching the bracketed field names
         (without the brackets). For example, if [First Name] is in the PDF,
         the Excel should have a column header "First Name"
       - Config file: Text file with the following format:
             excel_file = path/to/data.xlsx
             template = path/to/template.pdf
             output_directory = path/to/output
             filename_field1 = First Name  # Optional - uses timestamp if both fields omitted
             filename_field2 = Last Name   # Optional - uses timestamp if both fields omitted
    
    2. Run the script:
       python pdf_template_fill.py <config_file>
    
       Example:
       python pdf_template_fill.py config.txt

Note:
    - Field names in PDF must match Excel headers exactly (excluding brackets)
    - Field names are case-sensitive and space-sensitive: [First Name] ≠ [First_Name] ≠ [first name]
    - No automatic normalization of field names is performed
    - Output files will be named using the specified fields (or timestamp if omitted)
    - The script preserves all PDF formatting, images, and other content
"""

import sys
import os
import re
import datetime as dt
import time
from openpyxl import load_workbook
import traceback
import fitz  # PyMuPDF
from utils import format_date, sanitize_filename, read_config, get_unique_filename  # Import shared utilities

def find_fields_in_pdf(pdf_path):
    """
    Find all bracketed fields in the PDF document.
    
    Args:
        pdf_path: Path to the PDF file
        
    Returns:
        set: Set of unique field names found (without brackets)
    """
    fields = set()
    pattern = r'\[([^\]]+)\]'
    
    try:
        # Open PDF
        with fitz.open(pdf_path) as doc:
            # Search each page
            for page in doc:
                # Get text from page
                text = page.get_text()
                # Find all matches
                matches = re.finditer(pattern, text)
                fields.update(match.group(1) for match in matches)
    
    except Exception as e:
        print(f"Error reading PDF: {e}")
        raise
        
    return fields

def replace_fields_in_pdf(pdf_path, output_path, data):
    """
    Replace all bracketed fields with corresponding values.
    Attempts to match original text properties (font, size, color).
    
    Args:
        pdf_path: Path to the template PDF
        output_path: Path where to save the modified PDF
        data (dict): Dictionary of field names and their values
    """
    try:
        # Create a mapping of field names to values - only do this once per template
        if not hasattr(replace_fields_in_pdf, 'field_mapping'):
            replace_fields_in_pdf.field_mapping = {}
            replace_fields_in_pdf.font_substitutions = {}
            
            # Common font name mappings from Word/PDF to standard PostScript names
            replace_fields_in_pdf.font_mappings = {
                'TimesNewRomanPSMT': 'Times-Roman',
                'TimesNewRomanPS': 'Times-Roman',
                'TimesNewRoman': 'Times-Roman',
                'ArialMT': 'Arial',
                'ArialMS': 'Arial',
                'Calibri': 'Helvetica',
                'CalibriLight': 'Helvetica',
                'Cambria': 'Times-Roman',
                'Georgia': 'Times-Roman',
                'SegoeUI': 'Helvetica',
                'Verdana': 'Helvetica',
                'Symbol': 'Symbol',
                'ZapfDingbats': 'ZapfDingbats'
            }
            
            # Add a simple field mapping with exact field names
            for key, value in data.items():
                replace_fields_in_pdf.field_mapping[f"[{key}]"] = value
        else:
            # Update values in existing mapping
            for key, value in data.items():
                replace_fields_in_pdf.field_mapping[f"[{key}]"] = value
        
        # Open the PDF
        doc = fitz.open(pdf_path)
        
        # Track replacements for verification
        replacements_made = 0
        
        # Define fallback fonts in order of preference
        fallback_fonts = ['Helvetica', 'Arial', 'Times-Roman']
        
        # Process each page
        for page_num, page in enumerate(doc):
            # Search for each field directly
            for field, value in replace_fields_in_pdf.field_mapping.items():
                # Find all instances of this field on the page
                field_instances = page.search_for(field)
                
                for inst in field_instances:
                    # Get text properties for this instance
                    spans = page.get_text("dict", clip=inst)["blocks"]
                    if not spans:
                        continue
                    
                    try:
                        # Extract text properties
                        span = spans[0]["lines"][0]["spans"][0]
                        original_font = span.get("font", "Helvetica")
                        font_size = span.get("size", 11)
                        color = span.get("color", (0, 0, 0))
                    except (IndexError, KeyError):
                        # Fallback values
                        original_font = "Helvetica"
                        font_size = 11
                        color = (0, 0, 0)
                    
                    # Check if we've already found a substitution for this font
                    if original_font in replace_fields_in_pdf.font_substitutions:
                        font_name = replace_fields_in_pdf.font_substitutions[original_font]
                    else:
                        # Check if there's a mapping for this font
                        mapped_font = replace_fields_in_pdf.font_mappings.get(original_font)
                        if mapped_font:
                            try:
                                fitz.get_text_length(value, fontname=mapped_font, fontsize=font_size)
                                font_name = mapped_font
                                replace_fields_in_pdf.font_substitutions[original_font] = mapped_font
                            except ValueError:
                                font_name = None
                        else:
                            font_name = None
                            
                        # If no mapping worked, try the original font
                        if font_name is None:
                            try:
                                fitz.get_text_length(value, fontname=original_font, fontsize=font_size)
                                font_name = original_font
                                replace_fields_in_pdf.font_substitutions[original_font] = original_font
                            except ValueError:
                                # Try fallback fonts
                                for fallback_font in fallback_fonts:
                                    try:
                                        fitz.get_text_length(value, fontname=fallback_font, fontsize=font_size)
                                        font_name = fallback_font
                                        print(f"Using fallback font '{fallback_font}' instead of '{original_font}'")
                                        replace_fields_in_pdf.font_substitutions[original_font] = fallback_font
                                        break
                                    except ValueError:
                                        continue
                                
                                if font_name is None:  # If no fallback worked
                                    font_name = "Helvetica"
                                    print(f"Warning: Using Helvetica as fallback for '{original_font}'")
                                    replace_fields_in_pdf.font_substitutions[original_font] = "Helvetica"
                    
                    # Create redaction annotation to completely remove the original text
                    redact = page.add_redact_annot(inst)
                    page.apply_redactions()
                    
                    # Insert the new text at the original position
                    # Add a small padding to x position to prevent text from touching the edges
                    padding = font_size * 0.2  # 20% of font size as padding
                    # Use y1 (bottom) coordinate and offset up slightly for proper baseline alignment
                    baseline_offset = font_size * 0.2  # Offset up by 20% of font size
                    page.insert_text(
                        point=(inst.x0 + padding, inst.y1 - baseline_offset),
                        text=value,
                        fontname=font_name,
                        fontsize=font_size,
                        color=color
                    )
                    
                    replacements_made += 1
        
        if replacements_made == 0:
            print("Warning: No replacements were made. Check if field names match exactly.")
        else:
            print(f"Successfully made {replacements_made} replacements")
        
        # Save the modified PDF
        doc.save(output_path, garbage=4, deflate=True, clean=True)
        doc.close()
        
        # Verify the output file exists and is not empty
        if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
            raise Exception("Failed to create output PDF or file is empty")
        
    except Exception as e:
        print(f"Error modifying PDF: {e}")
        raise

def main():
    """Main function to process PDF documents."""
    if len(sys.argv) != 2:
        print("Usage: python pdf_template_filler.py <config_file>")
        sys.exit(1)
    
    try:
        total_start_time = time.time()
        config = read_config(sys.argv[1])
        
        # Extract configuration
        excel_file = config['excel_file']
        pdf_template = config['template']
        output_directory = config['output_directory']
        filename_field1 = config.get('filename_field1', '')
        filename_field2 = config.get('filename_field2', '')
        
        # Create output directory if it doesn't exist
        os.makedirs(output_directory, exist_ok=True)
        
        # Find fields in PDF template - do this once before processing rows
        template_fields = find_fields_in_pdf(pdf_template)
        print(f"\nFound {len(template_fields)} unique fields in PDF template:")
        print(", ".join(sorted(template_fields)))
        
        # Read Excel data
        wb = load_workbook(filename=excel_file, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        # Verify all template fields exist in Excel headers - do this once before processing rows
        missing_fields = []
        for field in template_fields:
            if field not in headers:
                missing_fields.append(field)
        
        if missing_fields:
            raise ValueError(f"Fields in PDF template not found in Excel headers: {', '.join(missing_fields)}")
        
        # Verify filename fields exist in headers if specified - do this once before processing rows
        if filename_field1 and filename_field1 not in headers:
            raise ValueError(f"Specified filename field '{filename_field1}' not found in Excel headers")
        
        if filename_field2 and filename_field2 not in headers:
            raise ValueError(f"Specified filename field '{filename_field2}' not found in Excel headers")
        
        if not filename_field1 and not filename_field2:
            print("No filename fields specified - using timestamps for output files")
        
        # Count total non-empty rows
        total_files = sum(1 for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row))
        processed_count = 0
        success_count = 0
        
        # Process each row
        for row_cells in ws.iter_rows(min_row=2):
            row = [cell.value for cell in row_cells]
            if not any(row):  # Skip empty rows
                continue
            
            processed_count += 1
            start_time = time.time()
            
            try:
                # Create data dictionary with formatted dates
                data = {}
                for i, cell in enumerate(row_cells):
                    if i < len(headers):
                        header = headers[i]
                        value = cell.value
                        
                        # Format dates consistently
                        if isinstance(value, (dt.datetime, dt.date)):
                            value = format_date(value)
                        elif value is not None:
                            value = str(value)
                        else:
                            value = ''
                            
                        data[header] = value
                
                # Generate output filename
                if filename_field1 or filename_field2:
                    field1_value = data.get(filename_field1, '').strip()
                    field2_value = data.get(filename_field2, '').strip()
                    filename = f"{field1_value} {field2_value}".strip()
                else:
                    filename = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # Sanitize filename
                filename = sanitize_filename(filename)
                
                # Create output path and handle duplicates
                base_path = os.path.join(output_directory, filename)
                output_path = get_unique_filename(base_path, "pdf")
                
                # Replace fields and save PDF
                replace_fields_in_pdf(pdf_template, output_path, data)
                
                success_count += 1
                elapsed_time = time.time() - start_time
                print(f"Processed {processed_count}/{total_files}: {os.path.basename(output_path)} in {elapsed_time:.1f} seconds")
                
            except Exception as e:
                print(f"Error processing row {processed_count}: {str(e)}")
                traceback.print_exc()
        
        # Print summary
        total_time = time.time() - total_start_time
        print("\nProcessing Summary:")
        print(f"Total files processed: {success_count}/{total_files}")
        print(f"Total processing time: {total_time:.1f} seconds")
        print(f"Average time per file: {(total_time/total_files):.1f} seconds")
        print(f"Output directory: {os.path.abspath(output_directory)}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main() 