"""
PDF Form Filler

This script uses pdfrw to fill PDF forms and PyMuPDF (fitz) to flatten form fields.

Requirements:
    - Python 3.6 or higher
    - Required packages (install exact versions to ensure compatibility):
        pip install openpyxl==3.0.10     # For Excel file handling
        pip install pdfrw==0.4.0         # For PDF form filling
        pip install PyMuPDF==1.21.1      # For field flattening

Usage:
    1. Prepare your files:
       - Excel file: First row should contain headers that match PDF form field names
       - PDF template: Should have fillable form fields
       - Config file: Text file with the following format 
         (the filename fields specify how each output PDF is named):
            excel_file = path/to/data.xlsx
            template = path/to/template.pdf
            output_directory = path/to/output
            filename_field1 = First Name  # Optional - uses timestamp if both fields omitted
            filename_field2 = Last Name   # Optional - uses timestamp if both fields omitted
            max_threads = 4               # Optional - number of concurrent processing threads (default: 4)
    
    2. Run the script:
       python pdf_form_filler.py <config_file>
    
       Example:
       python pdf_form_filler.py config.txt

Important Notes:
    - Troubleshoooting Tip: If having trouble with a particular field in the PDF form, try 
      deleting the field in Acrobat and re-adding it, rather than modifying it (moving it or 
      renaming it). For some reason, if you move or rename an existing form field after saving 
      it, the change is not reflected in the resulting generated PDF.
    - PDF form field names MUST NOT contain parentheses. The script will not work correctly
      with field names like "Name (Legal)" due to how PDF form fields are processed.
    - Excel headers must exactly match the PDF form field names for proper filling.
    - The multi-threading feature significantly improves processing speed for large batches.
"""

import sys
import os
from datetime import datetime
from openpyxl import load_workbook
from pdfrw import PdfReader, PdfWriter, PdfDict, PdfName, PdfObject
import fitz
import pymupdf
import time
import traceback
import re
import concurrent.futures
import threading
import signal

# Global flag to track if script should exit (for handling keyboard interrupts)
should_exit = False

# Thread-local storage for resource management
thread_local = threading.local()

# Set up signal handler for graceful shutdown
def signal_handler(sig, frame):
    """Handle keyboard interrupt and other signals to gracefully shut down."""
    global should_exit
    print("\nKeyboard interrupt detected. Shutting down gracefully...")
    print("Please wait for current tasks to finish (this may take a moment)...")
    
    # Set flag for threads to check
    should_exit = True
    
    # Don't exit immediately, let cleanup happen in main thread

# Register signal handlers
signal.signal(signal.SIGINT, signal_handler)  # Handles Ctrl+C
signal.signal(signal.SIGTERM, signal_handler)  # Handles termination signal

def normalize_field_name(name):
    """Normalize field name to handle variations in capitalization and spacing."""
    if not name:
        return []
    
    # Generate variations with different capitalizations and spacings
    name = str(name).strip()
    variations = [
        name,
        name.lower(),
        name.upper(),
        name.title(),
        name.replace(" ", ""),
        name.replace(" ", "_"),
        name.replace("_", " ")
    ]
    
    # Remove duplicates and empty strings
    return list(set(var for var in variations if var))

def read_excel_data(excel_path):
    """Read data from Excel file."""
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        sheet = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Read all rows (excluding header)
        data = []
        for row in list(sheet.rows)[1:]:
            row_data = {headers[i]: cell.value 
                       for i, cell in enumerate(row) 
                       if i < len(headers)}
            data.append(row_data)
            
        return headers, data
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None, None

def fill_pdf_form(template_path, data_row, temp_output_path):
    """Fill PDF form using pdfrw."""
    if should_exit:
        return False
        
    try:
        # Read the PDF template
        template = PdfReader(template_path)
        
        # Create PDF writer
        writer = PdfWriter()
        
        # Check if PDF has form fields
        if not template.Root.AcroForm:
            print("Error: No form fields found in PDF template")
            return False
            
        # Get form fields
        form_fields = template.Root.AcroForm.Fields
        if not form_fields:
            print("Warning: No form fields found in PDF")
        
        # Update form fields
        fields_filled = 0
        for field in form_fields:
            if field.T:
                key = field.T[1:-1]  # Remove parentheses from field name
                if key in data_row:
                    value = data_row[key]
                    if value is None or str(value).strip() == '':
                        field.V = ''
                    else:
                        field.V = str(value).strip()
                    field.AP = ''
                    fields_filled += 1
                
        # Set form flags
        template.Root.AcroForm.update(PdfDict(
            NeedAppearances=PdfObject('true')
        ))
        
        # Add all pages to the writer
        for page in template.pages:
            writer.addpage(page)
        
        # Write the filled PDF
        writer.write(temp_output_path)
        return True
        
    except Exception as e:
        print(f"Error filling PDF form: {e}")
        return False

def flatten_fields(input_path, output_path, fields_to_flatten):
    """Flatten specified fields using PyMuPDF."""
    if should_exit:
        return False
        
    pymupdf.TOOLS.mupdf_display_errors(False)
    doc = None
    success = False
    try:
        # Open the filled PDF
        doc = fitz.open(input_path)
        
        # Process each page
        for page_num, page in enumerate(doc):
            if should_exit:
                return False
                
            try:
                # Get form fields (widgets) on the page
                fields = list(page.widgets())
                
                for field in fields:
                    if should_exit:
                        return False
                        
                    try:
                        # Get field name safely
                        field_name = getattr(field, 'field_name', None)
                        if not field_name:
                            continue
                            
                        if field_name in fields_to_flatten:
                            
                            # Get field value safely
                            try:
                                field_value = field.field_value
                            except:
                                try:
                                    # Alternative way to get value
                                    field_value = field.text
                                except:
                                    print(f"Could not get value for field {field_name}")
                                    field_value = None
                            
                            # Check if field has a value
                            if field_value and str(field_value).strip():
                                try:
                                    # Get field rectangle
                                    rect = field.rect
                                    
                                    # Try to get font size from field
                                    try:
                                        font_size = field.font_size
                                        if not font_size or font_size <= 0:
                                            font_size = 12
                                    except:
                                        font_size = 12
                                                                        
                                    # Calculate text position (slightly offset from top-left)
                                    x = rect.x0 + 2  # 2 point offset from left
                                    y = rect.y0 + font_size  # offset by font size from top
                                    
                                    # Insert text at the calculated position
                                    page.insert_text(
                                        point=(x, y),
                                        text=str(field_value),
                                        fontsize=font_size,
                                        color=(0, 0, 0)  # Black text
                                    )
                                    
                                    # Try to remove the form field
                                    try:
                                        page.delete_widget(field)
                                    except Exception as e:
                                        print(f"Warning: Could not delete widget for {field_name}: {e}")
                                        
                                except Exception as e:
                                    print(f"Error processing field '{field_name}': {e}")
                                    continue
                            else:
                                # Field is empty, just remove the widget
                                try:
                                    page.delete_widget(field)
                                except Exception as e:
                                    print(f"Warning: Could not remove empty field {field_name}: {e}")
                                    
                    except Exception as field_error:
                        print(f"Error processing field: {field_error}")
                        continue
                        
            except Exception as page_error:
                print(f"Error processing page {page_num + 1}: {page_error}")
                continue
        
        # Save the modified PDF
        doc.save(output_path, garbage=4, deflate=True, clean=True)
        
        # Verify the file was saved
        success = os.path.exists(output_path) and os.path.getsize(output_path) > 0
        if not success:
            print("Error: Failed to save PDF or file is empty")
            
        return success
            
    except Exception as e:
        print(f"Error flattening PDF: {e}")
        return False
    finally:
        if doc:
            try:
                doc.close()
                doc = None
                import gc
                gc.collect()  # Force garbage collection
            except:
                pass

def process_pdf(template_path, data_row, output_path, fields_to_flatten):
    """Process a single PDF form - fill and flatten."""
    if should_exit:
        return False
        
    temp_path = output_path + '.temp.pdf'
    try:
        # Step 1: Fill the form using pdfrw
        if not fill_pdf_form(template_path, data_row, temp_path):
            print("Failed to fill PDF form")
            return False
        
        # Step 2: Flatten specified fields using PyMuPDF
        if fields_to_flatten:
            if not flatten_fields(temp_path, output_path, fields_to_flatten):
                print("Failed to flatten fields")
                return False
            # Force garbage collection after flattening
            import gc
            gc.collect()
        else:
            # If no fields to flatten, just rename the temp file
            try:
                os.replace(temp_path, output_path)
            except Exception as e:
                print(f"Error moving file: {e}")
                return False
        
        return True
    except Exception as e:
        print(f"Error processing PDF: {e}")
        return False
    finally:
        # Always try to remove the temp file with retries
        if os.path.exists(temp_path):
            for delay in [0, 0.2, 0.5, 1.0]:  # Longer delays
                try:
                    time.sleep(delay)
                    os.remove(temp_path)
                    break  # Success, exit the retry loop
                except Exception as e:
                    if delay == 1.0:  # Only print warning on last attempt
                        print(f"Warning: Could not remove temporary file {temp_path}: {e}")

def sanitize_filename(filename):
    """Remove invalid characters from filename."""
    # Replace invalid characters with underscores
    sanitized = re.sub(r'[\\/*?:"<>|]', "_", filename)
    # Remove leading/trailing spaces and dots
    sanitized = sanitized.strip(". ")
    # Default filename if empty
    if not sanitized:
        sanitized = "document"
    return sanitized

def read_config(config_path):
    """
    Read configuration from a file.
    
    Args:
        config_path (str): Path to the configuration file
        
    Returns:
        dict: Configuration parameters
        
    Raises:
        ValueError: If required fields are missing
    """
    config = {}
    required_fields = ['excel_file', 'template', 'output_directory']
    
    try:
        with open(config_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    key, value = map(str.strip, line.split('=', 1))
                    config[key] = value
    except Exception as e:
        raise ValueError(f"Error reading config file: {str(e)}")
    
    # Check for required fields
    missing = [field for field in required_fields if field not in config]
    if missing:
        raise ValueError(f"Missing required fields in config file: {', '.join(missing)}")
    
    return config

def process_row_task(args):
    """Process a single row from the Excel file (for threading)."""
    global should_exit
    if should_exit:
        return None
        
    row_idx, template_path, data, output_dir, headers, filename_field1, filename_field2, total_files = args
    
    start_time = time.time()
    idx = row_idx + 1  # Human-readable index (1-based)
    
    try:
        # Generate output filename from specified fields
        # Safely handle non-string values by converting to string first
        field1_value = ''
        if filename_field1:
            value = data.get(filename_field1, '')
            field1_value = str(value).strip() if value is not None else ''
            
        field2_value = ''
        if filename_field2:
            value = data.get(filename_field2, '')
            field2_value = str(value).strip() if value is not None else ''
        
        if field1_value or field2_value:
            filename = f"{field1_value} {field2_value}".strip()
        else:
            # Fallback to timestamp if both fields are empty
            filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{idx}"
        
        # Sanitize filename
        filename = sanitize_filename(filename)
        
        # Add .pdf extension and handle duplicates
        output_path = os.path.join(output_dir, f"{filename}.pdf")
        counter = 1
        while os.path.exists(output_path):
            output_path = os.path.join(output_dir, f"{filename}_{counter}.pdf")
            counter += 1
        
        success = process_pdf(template_path, data, output_path, headers)
        elapsed_time = time.time() - start_time
        
        # Prepare result with progress info included
        result = {
            'success': success,
            'output_path': output_path,
            'index': idx,
            'total': total_files,
            'elapsed_time': elapsed_time
        }
        
        return result
    except Exception as e:
        print(f"Error processing row {idx}: {str(e)}")
        return {
            'success': False,
            'index': idx,
            'total': total_files,
            'error': str(e)
        }

def main():
    """Main function to process PDF forms."""
    global should_exit
    
    if len(sys.argv) != 2:
        print("Usage: python pdf_form_filler.py <config_file>")
        sys.exit(1)
    
    try:
        total_start_time = time.time()
        config = read_config(sys.argv[1])
        
        # Extract configuration
        excel_file = config['excel_file']
        pdf_template = config['template']
        output_directory = config['output_directory']
        filename_field1 = config.get('filename_field1', '')
        filename_field2 = config.get('filename_field2', '')
        
        # Get threads configuration
        try:
            max_threads = int(config.get('max_threads', '4'))
            if max_threads < 1:
                max_threads = 1
            print(f"Using {max_threads} processing threads")
        except ValueError:
            max_threads = 4
            print(f"Invalid max_threads value, using default: {max_threads}")
        
        # Create output directory if it doesn't exist
        os.makedirs(output_directory, exist_ok=True)
        
        # Read Excel data
        wb = load_workbook(filename=excel_file, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        # Verify filename fields exist in headers if specified
        if filename_field1:
            variations1 = normalize_field_name(filename_field1)
            if not any(var in headers for var in variations1):
                raise ValueError(f"Specified filename field '{filename_field1}' not found in Excel headers")
            filename_field1 = next(var for var in variations1 if var in headers)
            
        if filename_field2:
            variations2 = normalize_field_name(filename_field2)
            if not any(var in headers for var in variations2):
                raise ValueError(f"Specified filename field '{filename_field2}' not found in Excel headers")
            filename_field2 = next(var for var in variations2 if var in headers)
        
        if not filename_field1 and not filename_field2:
            print("No filename fields specified - using timestamps for output files")
        
        print(f"Found {len(headers)} fields in Excel headers")
        
        # Collect all rows to process (non-empty rows)
        rows_to_process = []
        for idx, row_cells in enumerate(ws.iter_rows(min_row=2)):
            if not any(cell.value for cell in row_cells):  # Skip empty rows
                continue
                
            # Create data dictionary
            data = {headers[i]: cell.value for i, cell in enumerate(row_cells) if i < len(headers)}
            rows_to_process.append((idx, data))
        
        total_files = len(rows_to_process)
        print(f"Found {total_files} non-empty rows to process")
        
        # Process rows in parallel using thread pool
        success_count = 0
        tasks = []
        
        print("\nStarting PDF form filling process...")
        print(f"Template: {pdf_template}")
        print(f"Output directory: {output_directory}")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_threads) as executor:
            # Create tasks for each row
            for idx, row_data in rows_to_process:
                if should_exit:
                    break
                    
                # Create task arguments
                task_args = (
                    idx,
                    pdf_template,
                    row_data,
                    output_directory,
                    headers,
                    filename_field1,
                    filename_field2,
                    total_files
                )
                
                # Submit task to executor
                future = executor.submit(process_row_task, task_args)
                tasks.append(future)
            
            # Process results as they complete
            try:
                for future in concurrent.futures.as_completed(tasks):
                    if should_exit and not future.done():
                        continue
                        
                    result = future.result()
                    if result and result.get('success'):
                        success_count += 1
                        print(f"[{result['index']}/{result['total']}] Successfully processed: {os.path.basename(result['output_path'])} in {result['elapsed_time']:.1f} seconds")
                    elif result:
                        print(f"[{result['index']}/{result['total']}] Failed to process PDF")
            except KeyboardInterrupt:
                # Backup handler in case signal handler doesn't catch it
                should_exit = True
                print("\nKeyboard interrupt detected. Waiting for current tasks to finish...")
            
            # Handle cancellation if needed
            if should_exit:
                cancelled_count = 0
                for future in tasks:
                    if not future.done():
                        future.cancel()
                        cancelled_count += 1
                
                if cancelled_count > 0:
                    print(f"Cancelled {cancelled_count} pending tasks")
        
        # Calculate final stats
        total_time = time.time() - total_start_time
        
        # Print summary
        print("\nProcessing Summary:")
        if should_exit:
            print(f"Process interrupted by user")
            print(f"Completed files: {success_count}/{total_files}")
            print(f"Cancelled/unprocessed: {total_files - success_count}")
        else:
            print(f"Total files processed: {success_count}/{total_files}")
            if success_count > 0:
                print(f"Total processing time: {total_time:.1f} seconds")
                print(f"Average time per file: {(total_time/success_count):.1f} seconds")
                if success_count > 1:
                    print(f"Files per minute: {60 * success_count / total_time:.1f}")
            
        print(f"Output directory: {os.path.abspath(output_directory)}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        traceback.print_exc()
        sys.exit(1)
    finally:
        # Clean exit if interrupted
        if should_exit:
            print("Script terminated due to user interrupt")
            sys.exit(0)

if __name__ == '__main__':
    main() 