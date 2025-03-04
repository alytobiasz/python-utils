#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Word Template to PDF Converter

This script combines the functionality of docx_template_filler.py and docx_to_pdf.py.
It first replaces bracketed fields in a Word document with values from an Excel file,
then converts the resulting documents to PDF format.

Requirements:
    - Python 3.6 or higher
    - Required Python packages:
        pip install python-docx==0.8.11    # For Word document handling
        pip install openpyxl==3.0.10       # For Excel file handling
    - Microsoft Word must be installed (Windows or macOS)
    - Python packages for PDF conversion:
        Windows: pip install pywin32
        macOS: pip install pyobjc

Usage:
    1. Prepare your files:
       - Word template: Use bracketed fields like [Field Name] where you want replacements
       - Excel file: First row should contain headers matching the bracketed field names
         (without the brackets). For example, if [First Name] is in the Word doc,
         the Excel should have a column header "First Name"
       - Config file: Text file with the following format:
            excel_file = path/to/data.xlsx
            template = path/to/template.docx
            output_directory = path/to/output
            filename_field1 = First Name  # Optional - uses timestamp if both fields omitted
            filename_field2 = Last Name   # Optional - uses timestamp if both fields omitted
            keep_word_file = false        # Optional - set to true to keep both .docx and .pdf
    
    2. Run the script:
       python word_template_to_pdf.py <config_file>
    
       Example:
       python word_template_to_pdf.py config.txt

Note:
    - Field names in Word doc must match Excel headers exactly (excluding brackets)
    - Fields are case-sensitive: [First_Name] ≠ [first_name]
    - Output files will be named using the specified fields (or timestamp if omitted)
"""

import sys
import os
import platform
from datetime import datetime
import time
import traceback

# Import functions from docx_template_filler.py
from docx_template_filler import (
    normalize_field_name,
    find_fields_in_document,
    replace_fields_in_document,
    read_config,
    sanitize_filename
)

# Import functions from docx_to_pdf.py
if platform.system() == 'Windows':
    from docx_to_pdf import convert_to_pdf_windows as convert_to_pdf
elif platform.system() == 'Darwin':  # macOS
    from docx_to_pdf import convert_to_pdf_macos as convert_to_pdf
else:
    print("Error: This script only supports Windows and macOS")
    sys.exit(1)

# Additional imports needed for document processing
from docx import Document
from openpyxl import load_workbook

def process_documents(config):
    """
    Process Word documents: fill templates with data and convert to PDF.
    
    Args:
        config (dict): Configuration parameters
    """
    try:
        total_start_time = time.time()
        
        # Extract configuration
        excel_file = config['excel_file']
        word_template = config['template']
        output_directory = config['output_directory']
        filename_field1 = config.get('filename_field1', '')
        filename_field2 = config.get('filename_field2', '')
        keep_word = config.get('keep_word_file', '').lower() == 'true'
        
        # Create output directory if it doesn't exist
        os.makedirs(output_directory, exist_ok=True)
        
        # Create pdf_exports subdirectory
        pdf_dir = os.path.join(output_directory, 'pdf_exports')
        os.makedirs(pdf_dir, exist_ok=True)
        
        # Load the template to find fields
        template_doc = Document(word_template)
        template_fields = find_fields_in_document(template_doc)
        print(f"\nFound {len(template_fields)} unique fields in Word template:")
        print(", ".join(sorted(template_fields)))
        
        # Read Excel data
        wb = load_workbook(filename=excel_file, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        # Verify all template fields exist in Excel headers
        missing_fields = []
        for field in template_fields:
            field_variations = normalize_field_name(field)
            if not any(var in headers for var in field_variations):
                missing_fields.append(field)
        
        if missing_fields:
            raise ValueError(f"Fields in Word template not found in Excel headers: {', '.join(missing_fields)}")
        
        # Verify filename fields exist in headers if specified
        if filename_field1:
            variations1 = normalize_field_name(filename_field1)
            if not any(var in headers for var in variations1):
                raise ValueError(f"Specified filename field '{filename_field1}' not found in Excel headers")
            filename_field1 = next(var for var in variations1 if var in headers)
            
        if filename_field2:
            variations2 = normalize_field_name(filename_field2)
            if not any(var in headers for var in variations2):
                raise ValueError(f"Specified filename field '{filename_field2}' not found in Excel headers")
            filename_field2 = next(var for var in variations2 if var in headers)
        
        if not filename_field1 and not filename_field2:
            print("No filename fields specified - using timestamps for output files")
        
        # Count total non-empty rows
        total_files = sum(1 for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row))
        processed_count = 0
        docx_success = 0
        pdf_success = 0
        docx_files = []  # Track created Word files for PDF conversion
        
        print("\nStep 1: Generating Word documents...")
        
        # Process each row to create Word documents
        for row_cells in ws.iter_rows(min_row=2):
            row = [cell.value for cell in row_cells]
            if not any(row):  # Skip empty rows
                continue
            
            processed_count += 1
            start_time = time.time()
            
            try:
                # Create data dictionary
                data = {headers[i]: str(val) if val is not None else '' 
                       for i, val in enumerate(row)}
                
                # Generate output filename
                if filename_field1 or filename_field2:
                    field1_value = data.get(filename_field1, '').strip()
                    field2_value = data.get(filename_field2, '').strip()
                    filename = f"{field1_value} {field2_value}".strip()
                else:
                    filename = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # Sanitize filename
                filename = sanitize_filename(filename)
                
                # Create output path
                docx_path = os.path.join(output_directory, f"{filename}.docx")
                
                # Handle duplicate filenames
                counter = 1
                while os.path.exists(docx_path):
                    filename = f"{filename}_{counter}"
                    docx_path = os.path.join(output_directory, f"{filename}.docx")
                    counter += 1
                
                # Create and save the filled document
                doc = Document(word_template)
                replace_fields_in_document(doc, data)
                doc.save(docx_path)
                
                docx_success += 1
                docx_files.append((docx_path, filename))
                elapsed_time = time.time() - start_time
                print(f"Generated {processed_count}/{total_files}: {filename}.docx in {elapsed_time:.1f} seconds")
                
            except Exception as e:
                print(f"Error processing row {processed_count}: {str(e)}")
                traceback.print_exc()
        
        print("\nStep 2: Converting to PDF...")
        
        # Convert Word documents to PDF
        for docx_path, filename in docx_files:
            try:
                success, message = convert_to_pdf(docx_path, pdf_dir)
                if success:
                    pdf_success += 1
                    if not keep_word:
                        try:
                            os.remove(docx_path)
                        except Exception as e:
                            print(f"Warning: Could not remove Word file {filename}.docx: {e}")
                print(message)
                
            except Exception as e:
                print(f"Error converting {filename}.docx to PDF: {e}")
                traceback.print_exc()
        
        # Print summary
        total_time = time.time() - total_start_time
        print("\nProcessing Summary:")
        print(f"Word documents generated: {docx_success}/{total_files}")
        print(f"PDF files created: {pdf_success}/{docx_success}")
        print(f"Total processing time: {total_time:.1f} seconds")
        print(f"Average time per file: {(total_time/total_files):.1f} seconds")
        print(f"Word files directory: {os.path.abspath(output_directory)}")
        print(f"PDF files directory: {os.path.abspath(pdf_dir)}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        traceback.print_exc()
        sys.exit(1)

def main():
    """Main function to handle command line arguments and start processing."""
    if len(sys.argv) != 2:
        print("Usage: python word_template_to_pdf.py <config_file>")
        sys.exit(1)
    
    # Check if running on supported OS
    if platform.system() not in ['Windows', 'Darwin']:
        print("Error: This script only supports Windows and macOS")
        sys.exit(1)
    
    config = read_config(sys.argv[1])
    process_documents(config)

if __name__ == "__main__":
    main() 